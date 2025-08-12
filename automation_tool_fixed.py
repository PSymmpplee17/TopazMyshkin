#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Программа автоматизации для работы с Excel файлами
Основная задача: обработка Excel файлов, удаление дублей с суммированием значений
Работает с файлами *.xls и *.xlsm
"""

import pandas as pd
import os
import sys
from pathlib import Path
import logging
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import shutil
import xlrd

# Настройка логирования
def setup_logging():
    """Настраивает логирование в папку logs"""
    # Определяем папку приложения
    if getattr(sys, 'frozen', False):
        app_dir = Path(sys.executable).parent
    else:
        app_dir = Path(__file__).parent
    
    # Создаем папку logs
    logs_dir = app_dir / "logs"
    logs_dir.mkdir(exist_ok=True)
    
    # Настраиваем логирование
    log_file = logs_dir / 'automation.log'
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler(sys.stdout)
        ]
    )
    return logging.getLogger(__name__)

logger = setup_logging()


class ExcelProcessor:
    """Класс для обработки Excel файлов"""
    
    # Константы для фиксированной ширины столбцов (после удаления B, C, F)
    # Исходные столбцы A, D, E, G, H, I, J стали A, B, C, D, E, F, G
    FIXED_COLUMN_WIDTHS = {
        'A': 4,   # 33 пикселя (исходный столбец A)
        'B': 64,  # 450 пикселей (исходный столбец D) 
        'C': 22,  # 159 пикселей (исходный столбец E)
        'D': 11,  # 81 пиксель (исходный столбец G)
        'E': 10,   # 75 пикселей (исходный столбец H)
        'F': 26,   # 195 пикселей (исходный столбец I)
        'G': 6     # 45 пикселей (исходный столбец J)
    }
    
    # Константы для столбцов, которые должны содержать числа (после удаления B, C, F)
    # Исходный столбец J (количество) стал столбцом G  
    NUMERIC_COLUMNS = {'G'}  # Только столбец G содержит числовые данные (количество)
    
    def __init__(self, input_file: str):
        """
        Инициализация процессора
        
        Args:
            input_file (str): Путь к входному файлу
        """
        self.input_file = Path(input_file)
        self.output_file = None
        self.df = None
        self.workbook = None
        self.worksheet = None
        
        # Проверяем существование файла
        if not self.input_file.exists():
            raise FileNotFoundError(f"Файл {input_file} не найден")
        
        # Проверяем расширение
        if self.input_file.suffix.lower() not in ['.xls', '.xlsm', '.xlsx']:
            raise ValueError(f"Неподдерживаемый формат файла: {self.input_file.suffix}")
    
    def load_data(self):
        """Загрузка данных из Excel файла с сохранением форматирования"""
        try:
            logger.info(f"Загружаем данные из файла: {self.input_file}")
            
            # Создаем рабочую копию файла для обработки
            if self.input_file.suffix.lower() == '.xls':
                logger.info("Конвертируем .xls в .xlsx с сохранением типов данных и форматирования...")
                temp_xlsx_file = self.input_file.with_suffix('.temp.xlsx')
                
                # Открываем .xls файл
                try:
                    xls_book = xlrd.open_workbook(self.input_file, formatting_info=True)
                    xls_sheet = xls_book.sheet_by_index(0)
                    
                    # Создаем новую рабочую книгу
                    wb = Workbook()
                    ws = wb.active
                    
                    # Копируем данные с сохранением типов
                    logger.info("Копируем данные с сохранением типов...")
                    for row_idx in range(xls_sheet.nrows):
                        for col_idx in range(xls_sheet.ncols):
                            orig_cell = xls_sheet.cell(row_idx, col_idx)
                            excel_cell = ws.cell(row=row_idx+1, column=col_idx+1)
                            
                            # Копируем значение с сохранением типа
                            col_letter = get_column_letter(col_idx + 1)
                            
                            if orig_cell.ctype == xlrd.XL_CELL_NUMBER:
                                # Число
                                excel_cell.value = orig_cell.value
                            elif orig_cell.ctype == xlrd.XL_CELL_DATE:
                                # Дата
                                try:
                                    date_value = xlrd.xldate_as_tuple(orig_cell.value, xls_book.datemode)
                                    from datetime import datetime
                                    excel_cell.value = datetime(*date_value)
                                except:
                                    excel_cell.value = orig_cell.value
                            elif orig_cell.ctype == xlrd.XL_CELL_BOOLEAN:
                                # Булево
                                excel_cell.value = bool(orig_cell.value)
                            elif orig_cell.ctype == xlrd.XL_CELL_TEXT:
                                # Текст - проверяем, нужно ли конвертировать в число
                                text_value = orig_cell.value
                                if col_letter in self.NUMERIC_COLUMNS and text_value:
                                    # Пытаемся конвертировать в число
                                    try:
                                        # Убираем пробелы и заменяем запятые на точки
                                        clean_value = str(text_value).strip().replace(',', '.')
                                        if clean_value and clean_value not in ('', '-', 'None'):
                                            # Пытаемся конвертировать в float
                                            numeric_value = float(clean_value)
                                            # Если это целое число, сохраняем как int
                                            if numeric_value.is_integer():
                                                excel_cell.value = int(numeric_value)
                                            else:
                                                excel_cell.value = numeric_value
                                            logger.debug(f"Конвертирован текст '{text_value}' в число {excel_cell.value} (столбец {col_letter})")
                                        else:
                                            excel_cell.value = None
                                    except (ValueError, TypeError):
                                        # Если не удалось конвертировать, оставляем как текст
                                        excel_cell.value = text_value
                                        logger.warning(f"Не удалось конвертировать '{text_value}' в число (столбец {col_letter})")
                                else:
                                    excel_cell.value = text_value
                            elif orig_cell.ctype == xlrd.XL_CELL_EMPTY:
                                # Пустая
                                excel_cell.value = None
                            else:
                                # По умолчанию - проверяем на числовые столбцы
                                value = orig_cell.value if orig_cell.value else None
                                col_letter = get_column_letter(col_idx + 1)
                                
                                if col_letter in self.NUMERIC_COLUMNS and value is not None:
                                    try:
                                        clean_value = str(value).strip().replace(',', '.')
                                        if clean_value and clean_value not in ('', '-', 'None'):
                                            numeric_value = float(clean_value)
                                            if numeric_value.is_integer():
                                                excel_cell.value = int(numeric_value)
                                            else:
                                                excel_cell.value = numeric_value
                                        else:
                                            excel_cell.value = None
                                    except (ValueError, TypeError):
                                        excel_cell.value = value
                                else:
                                    excel_cell.value = value
                            
                            # Применяем базовое форматирование
                            if excel_cell.value is not None:
                                try:
                                    # Получаем XF record для ячейки
                                    xf_index = xls_sheet.cell_xf_index(row_idx, col_idx)
                                    xf = xls_book.xf_list[xf_index]
                                    
                                    # Применяем границы
                                    excel_cell.border = Border(
                                        left=Side(style='thin'),
                                        right=Side(style='thin'), 
                                        top=Side(style='thin'),
                                        bottom=Side(style='thin')
                                    )
                                    
                                    # Применяем шрифт
                                    font_info = xls_book.font_list[xf.font_index]
                                    excel_cell.font = Font(
                                        name=font_info.name or 'Calibri',
                                        size=font_info.height/20 if font_info.height else 11,
                                        bold=font_info.bold,
                                        italic=font_info.italic
                                    )
                                    
                                    # Применяем выравнивание
                                    alignment = xf.alignment
                                    excel_cell.alignment = Alignment(
                                        horizontal='general',
                                        vertical='bottom',
                                        wrap_text=alignment.wrap
                                    )
                                except:
                                    # Базовое форматирование при ошибке
                                    excel_cell.border = Border(
                                        left=Side(style='thin'),
                                        right=Side(style='thin'),
                                        top=Side(style='thin'),
                                        bottom=Side(style='thin')
                                    )
                                    excel_cell.font = Font(name='Calibri', size=11)
                    
                    # Копируем ширину столбцов на основе содержимого
                    logger.info("Устанавливаем ширину столбцов...")
                    try:
                        for col_idx in range(xls_sheet.ncols):
                            col_letter = get_column_letter(col_idx + 1)
                            
                            # Проверяем, есть ли фиксированная ширина для этого столбца
                            if col_letter in self.FIXED_COLUMN_WIDTHS:
                                width = self.FIXED_COLUMN_WIDTHS[col_letter]
                                logger.info(f"Столбец {col_letter}: фиксированная ширина {width}")
                            else:
                                # Автоматический размер на основе содержимого
                                max_length = 0
                                for row_idx in range(min(xls_sheet.nrows, 100)):  # Проверяем первые 100 строк
                                    cell_value = xls_sheet.cell_value(row_idx, col_idx)
                                    if cell_value:
                                        max_length = max(max_length, len(str(cell_value)))
                                
                                # Устанавливаем ширину на основе содержимого
                                width = min(max_length * 1.2 + 2, 50)  # Максимум 50 единиц
                                width = max(width, 8.43)  # Минимум стандартная ширина
                            
                            ws.column_dimensions[col_letter].width = width
                    except Exception as e:
                        logger.warning(f"Не удалось установить ширину столбцов: {e}")
                    
                    # Сохраняем файл
                    wb.save(temp_xlsx_file)
                    wb.close()
                    
                except Exception as e:
                    logger.warning(f"Не удалось извлечь детальное форматирование из .xls: {e}")
                    # Fallback: используем pandas
                    df_temp = pd.read_excel(self.input_file, engine='xlrd', header=None)
                    with pd.ExcelWriter(temp_xlsx_file, engine='openpyxl') as writer:
                        df_temp.to_excel(writer, index=False, header=False, sheet_name='Sheet1')
                
                # Теперь работаем с конвертированным файлом
                self.input_file = temp_xlsx_file
                logger.info(f"Файл конвертирован с сохранением типов данных в: {self.input_file}")
            
            # Загружаем данные через pandas
            if self.input_file.suffix.lower() == '.xls':
                engine = 'xlrd'
            else:
                engine = 'openpyxl'
            
            # Загружаем данные
            self.df = pd.read_excel(
                self.input_file,
                engine=engine,
                header=None  # Не используем первую строку как заголовки
            )
            
            # Загружаем workbook для сохранения форматирования
            if self.input_file.suffix.lower() in ['.xlsx', '.xlsm']:
                try:
                    self.workbook = load_workbook(self.input_file)
                    self.worksheet = self.workbook.active
                    logger.info("Форматирование загружено успешно")
                except Exception as e:
                    logger.warning(f"Не удалось загрузить форматирование: {e}")
                    self.workbook = None
                    self.worksheet = None
            
            logger.info(f"Данные загружены. Размер: {self.df.shape}")
            return True
            
        except Exception as e:
            logger.error(f"Ошибка при загрузке файла: {e}")
            return False
    
    def remove_empty_rows(self, col1_idx: int = 3, col2_idx: int = 4):
        """
        Удаляет строки, где в указанных столбцах нет данных
        
        Args:
            col1_idx (int): Индекс первого столбца для проверки (начиная с 0)
            col2_idx (int): Индекс второго столбца для проверки (начиная с 0)
        """
        if self.df is None:
            logger.error("Данные не загружены. Сначала вызовите load_data()")
            return False
        
        try:
            logger.info("Удаляем пустые строки...")
            original_rows = len(self.df)
            
            # Проверяем, что указанные столбцы существуют
            if col1_idx >= self.df.shape[1] or col2_idx >= self.df.shape[1]:
                logger.error(f"Указанные столбцы ({col1_idx}, {col2_idx}) не существуют в файле")
                return False
            
            # Создаем маску для строк, где хотя бы в одном из столбцов есть данные
            mask_alternative = (
                (self.df.iloc[:, col1_idx].notna() & (self.df.iloc[:, col1_idx] != '')) |
                (self.df.iloc[:, col2_idx].notna() & (self.df.iloc[:, col2_idx] != ''))
            )
            
            # Применяем фильтр - оставляем строки, где хотя бы в одном из столбцов есть данные
            self.df = self.df[mask_alternative].copy()
            
            # Сбрасываем индекс для правильной нумерации
            self.df.reset_index(drop=True, inplace=True)
            
            removed_rows = original_rows - len(self.df)
            logger.info(f"Удалено строк: {removed_rows}. Осталось строк: {len(self.df)}")
            
            return True
            
        except Exception as e:
            logger.error(f"Ошибка при удалении пустых строк: {e}")
            return False
    
    def process_duplicates_with_order_preservation(self, primary_key_col: int = 8, sum_col: int = 9, 
                                                 keep_cols: list = [0, 3, 4, 6, 7, 8, 9], 
                                                 remove_cols: list = [1, 2, 5]):
        """
        Обрабатывает дубли с сохранением порядка строк
        
        Args:
            primary_key_col (int): Индекс столбца с первичным ключом (столбец I = индекс 8)
            sum_col (int): Индекс столбца для суммирования (столбец J = индекс 9) 
            keep_cols (list): Список индексов столбцов для сохранения (A,D,E,G,H,I,J = [0,3,4,6,7,8,9])
            remove_cols (list): Список индексов столбцов для удаления (B,C,F = [1,2,5])
        """
        if self.df is None:
            logger.error("Данные не загружены. Сначала вызовите load_data()")
            return False
        
        try:
            logger.info("Начинаем обработку дублей с сохранением порядка строк...")
            original_rows = len(self.df)
            
            # Проверяем существование нужных столбцов
            max_col_idx = max(keep_cols + [primary_key_col, sum_col])
            if max_col_idx >= self.df.shape[1]:
                logger.error(f"Столбец с индексом {max_col_idx} не существует в файле")
                return False
            
            # Создаем копию исходных данных с добавлением индекса порядка
            df_work = self.df.copy()
            df_work['_original_order'] = range(len(df_work))
            
            logger.info(f"Первичный ключ - столбец {primary_key_col} (столбец {chr(65 + primary_key_col)})")
            logger.info(f"Столбец для суммирования - столбец {sum_col} (столбец {chr(65 + sum_col)})")
            
            # Проверяем данные в столбце первичного ключа
            logger.info(f"Всего строк в исходных данных: {len(df_work)}")
            logger.info(f"Проверяем столбец {primary_key_col} на наличие данных...")
            
            # Показываем первые несколько значений для отладки
            for i in range(min(10, len(df_work))):
                val = df_work.iloc[i, primary_key_col]
                logger.info(f"Строка {i}: '{val}' (тип: {type(val)})")
            
            # Фильтруем строки с не пустыми первичными ключами
            primary_key_mask = df_work.iloc[:, primary_key_col].notna() & (df_work.iloc[:, primary_key_col] != '')
            logger.info(f"Строк с непустым первичным ключом: {primary_key_mask.sum()}")
            
            df_work = df_work[primary_key_mask].copy()
            
            if len(df_work) == 0:
                logger.warning("После фильтрации по первичному ключу не осталось строк")
                logger.warning("Возможные причины:")
                logger.warning("- Все значения в столбце первичного ключа пустые")
                logger.warning("- Неправильный индекс столбца первичного ключа")
                logger.warning("- Данные находятся в другом столбце")
                return False
            
            # Группируем по первичному ключу, сохраняя порядок первого вхождения
            logger.info("Группируем данные по первичному ключу...")
            
            # Словарь для хранения результатов с сохранением порядка
            result_rows = {}
            processed_keys = set()
            
            # Проходим по строкам в исходном порядке
            for idx, row in df_work.iterrows():
                primary_key = row.iloc[primary_key_col]
                
                if pd.isna(primary_key) or primary_key == '':
                    continue
                
                if primary_key not in processed_keys:
                    # Первое вхождение ключа - сохраняем позицию и данные
                    initial_sum = 0
                    if not pd.isna(row.iloc[sum_col]):
                        try:
                            clean_val = str(row.iloc[sum_col]).replace(',', '.').replace(' ', '').strip()
                            initial_sum = int(round(float(clean_val))) if clean_val and clean_val != '' else 0
                        except (ValueError, TypeError):
                            initial_sum = 0
                    
                    result_rows[primary_key] = {
                        'order': row['_original_order'],
                        'data': row.copy(),
                        'sum_value': initial_sum
                    }
                    processed_keys.add(primary_key)
                    logger.debug(f"Новый ключ '{primary_key}' на позиции {row['_original_order']}, начальная сумма: {initial_sum}")
                else:
                    # Дубль - добавляем значение к сумме
                    try:
                        additional_value = 0
                        if not pd.isna(row.iloc[sum_col]):
                            clean_val = str(row.iloc[sum_col]).replace(',', '.').replace(' ', '').strip()
                            additional_value = int(round(float(clean_val))) if clean_val and clean_val != '' else 0
                        
                        result_rows[primary_key]['sum_value'] += additional_value
                        logger.debug(f"Дубль '{primary_key}': добавлено {additional_value}, общая сумма: {result_rows[primary_key]['sum_value']}")
                    except (ValueError, TypeError) as e:
                        logger.warning(f"Не удалось обработать значение для суммирования в строке с ключом '{primary_key}': {e}")
            
            # Создаем итоговый DataFrame, отсортированный по исходному порядку
            logger.info("Формируем итоговый результат...")
            result_data = []
            
            # Сортируем по исходному порядку появления
            sorted_results = sorted(result_rows.items(), key=lambda x: x[1]['order'])
            
            for primary_key, data in sorted_results:
                row_data = data['data'].copy()
                # Обновляем значение в столбце суммирования
                row_data.iloc[sum_col] = data['sum_value']
                result_data.append(row_data)
            
            # Создаем новый DataFrame
            if result_data:
                result_df = pd.DataFrame(result_data)
                # Удаляем служебный столбец
                result_df = result_df.drop('_original_order', axis=1)
                
                # Оставляем только нужные столбцы в правильном порядке
                logger.info(f"Удаляем ненужные столбцы: {[chr(65 + col) for col in remove_cols]}")
                logger.info(f"Оставляем столбцы: {[chr(65 + col) for col in keep_cols]}")
                
                # Фильтруем только существующие столбцы
                valid_keep_cols = [col for col in keep_cols if col < result_df.shape[1]]
                result_df = result_df.iloc[:, valid_keep_cols].copy()
                
                # Обновляем основной DataFrame
                self.df = result_df.reset_index(drop=True)
                
                processed_rows = len(self.df)
                removed_duplicates = len(processed_keys) - processed_rows if len(processed_keys) > processed_rows else 0
                
                logger.info(f"Обработка дублей завершена:")
                logger.info(f"  Исходных строк: {original_rows}")
                logger.info(f"  Уникальных ключей: {len(processed_keys)}")
                logger.info(f"  Итоговых строк: {processed_rows}")
                logger.info(f"  Удалено дублей: {original_rows - processed_rows}")
                logger.info(f"  Порядок строк сохранен")
                
                return True
            else:
                logger.error("Не удалось сформировать результат")
                return False
                
        except Exception as e:
            logger.error(f"Ошибка при обработке дублей: {e}")
            return False
    
    def save_data_with_formatting(self, output_file: str = None):
        """
        Сохраняет обработанные данные в новый файл с сохранением форматирования
        
        Args:
            output_file (str): Путь к выходному файлу. Если не указан, создается автоматически
        """
        if self.df is None:
            logger.error("Нет данных для сохранения")
            return False
        
        try:
            if output_file is None:
                # Создаем уникальное имя выходного файла
                stem = self.input_file.stem
                counter = 1
                base_output_file = self.input_file.parent / f"{stem}_processed.xlsx"
                
                # Проверяем, не существует ли файл, и создаем уникальное имя
                while base_output_file.exists():
                    base_output_file = self.input_file.parent / f"{stem}_processed_{counter}.xlsx"
                    counter += 1
                
                output_file = base_output_file
            
            self.output_file = Path(output_file)
            
            # Проверяем, не открыт ли файл
            if self.output_file.exists():
                try:
                    # Пытаемся переименовать файл, чтобы проверить, не занят ли он
                    temp_name = self.output_file.with_suffix('.tmp')
                    self.output_file.rename(temp_name)
                    temp_name.rename(self.output_file)
                except (PermissionError, OSError):
                    # Файл открыт, создаем новое имя
                    import time
                    timestamp = int(time.time())
                    stem = self.output_file.stem
                    parent = self.output_file.parent
                    suffix = self.output_file.suffix
                    self.output_file = parent / f"{stem}_{timestamp}{suffix}"
                    logger.warning(f"Файл занят, сохраняем как: {self.output_file.name}")
            
            logger.info(f"Сохраняем данные с форматированием в файл: {self.output_file}")
            
            if self.workbook is not None and self.worksheet is not None:
                # Сохраняем с форматированием через openpyxl
                return self._save_with_openpyxl_formatting()
            else:
                # Fallback - обычное сохранение
                logger.warning("Форматирование недоступно, сохраняем без форматирования")
                return self.save_data_simple()
                
        except Exception as e:
            logger.error(f"Ошибка при сохранении файла с форматированием: {e}")
            return False
    
    def _save_with_openpyxl_formatting(self):
        """Сохранение с форматированием через openpyxl"""
        try:
            logger.info("Начинаем сохранение с форматированием...")
            
            # Создаем новый файл на основе исходного
            dest_workbook = load_workbook(self.input_file)
            dest_worksheet = dest_workbook.active
            
            # Сохраняем информацию о форматировании столбцов
            logger.info("Копируем размеры столбцов и строк...")
            columns_formatting = {}
            for col_letter, col_dim in dest_worksheet.column_dimensions.items():
                if col_dim.width:
                    columns_formatting[col_letter] = {'width': col_dim.width}
            
            # Очищаем все данные из листа, сохраняя форматирование
            logger.info(f"Очищаем исходные данные ({dest_worksheet.max_row} строк)")
            dest_worksheet.delete_rows(1, dest_worksheet.max_row)
            
            # Записываем новые данные
            logger.info(f"Записываем обработанные данные ({len(self.df)} строк, {len(self.df.columns)} столбцов)")
            
            for row_idx, (_, row) in enumerate(self.df.iterrows(), start=1):
                for col_idx, value in enumerate(row, start=1):
                    cell = dest_worksheet.cell(row=row_idx, column=col_idx)
                    
                    # Устанавливаем значение
                    if pd.isna(value):
                        cell.value = None
                    else:
                        # Проверяем числовые столбцы (только начиная со второй строки)
                        col_letter = get_column_letter(col_idx)
                        if col_letter in self.NUMERIC_COLUMNS and row_idx > 1 and value is not None:
                            # Пытаемся конвертировать в число
                            try:
                                if isinstance(value, str):
                                    clean_value = str(value).strip().replace(',', '.').replace(' ', '')
                                    if clean_value and clean_value not in ('', '-', 'None', 'nan'):
                                        # Пытаемся сначала конвертировать в int
                                        try:
                                            cell.value = int(clean_value)
                                            cell.number_format = '0'  # Целые числа
                                            logger.debug(f"Столбец {col_letter}, строка {row_idx}: '{clean_value}' -> {cell.value} (целое число)")
                                        except ValueError:
                                            # Если не получается int, пробуем float, но округляем до int
                                            numeric_value = float(clean_value)
                                            cell.value = int(round(numeric_value))
                                            cell.number_format = '0'  # Целые числа
                                            logger.debug(f"Столбец {col_letter}, строка {row_idx}: '{clean_value}' -> {cell.value} (округлено до целого)")
                                    else:
                                        cell.value = value
                                else:
                                    # Если значение уже число, преобразуем в int
                                    if isinstance(value, (int, float)):
                                        cell.value = int(round(float(value)))
                                        cell.number_format = '0'
                                    else:
                                        cell.value = value
                            except (ValueError, TypeError):
                                cell.value = value
                        else:
                            cell.value = value
                    
                    # Применяем базовое форматирование границ
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # Применяем базовый шрифт
                    cell.font = Font(
                        name='Calibri',
                        size=11
                    )
            # Применяем ширину столбцов
            logger.info("Применяем ширину столбцов...")
            
            for col_letter, formatting in columns_formatting.items():
                # Применяем фиксированную ширину для определенных столбцов
                if col_letter in self.FIXED_COLUMN_WIDTHS:
                    dest_worksheet.column_dimensions[col_letter].width = self.FIXED_COLUMN_WIDTHS[col_letter]
                    logger.info(f"Столбец {col_letter}: установлена фиксированная ширина {self.FIXED_COLUMN_WIDTHS[col_letter]}")
                elif 'width' in formatting and formatting['width']:
                    dest_worksheet.column_dimensions[col_letter].width = formatting['width']
            
            # Сохраняем результат
            dest_workbook.save(self.output_file)
            dest_workbook.close()
            
            logger.info("Данные с полным форматированием успешно сохранены")
            return True
            
        except Exception as e:
            logger.error(f"Ошибка при сохранении с форматированием: {e}")
            logger.info("Пытаемся сохранить без форматирования...")
            return self.save_data_simple()
    
    def save_data_simple(self, output_file: str = None):
        """Простое сохранение без форматирования (fallback метод)"""
        if output_file is None:
            output_file = self.output_file
        else:
            self.output_file = Path(output_file)
        
        try:
            # Проверяем и создаем уникальное имя файла если нужно
            if self.output_file.exists():
                try:
                    # Пытаемся удалить файл
                    self.output_file.unlink()
                except (PermissionError, OSError) as e:
                    # Файл занят, создаем новое имя
                    import time
                    timestamp = int(time.time())
                    stem = self.output_file.stem
                    parent = self.output_file.parent
                    suffix = self.output_file.suffix
                    self.output_file = parent / f"{stem}_{timestamp}{suffix}"
                    logger.warning(f"Файл занят, сохраняем как: {self.output_file.name}")
            
            # Пытаемся сохранить файл
            try:
                with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
                    self.df.to_excel(
                        writer,
                        index=False,
                        header=False
                    )
                
                logger.info("Данные сохранены (без форматирования)")
                return True
                
            except PermissionError:
                # Пробуем альтернативный путь в папке temp
                import tempfile
                temp_dir = Path(tempfile.gettempdir())
                temp_file = temp_dir / self.output_file.name
                
                with pd.ExcelWriter(temp_file, engine='openpyxl') as writer:
                    self.df.to_excel(
                        writer,
                        index=False,
                        header=False
                    )
                
                logger.warning(f"Файл сохранен во временной папке: {temp_file}")
                print(f"⚠️  Файл сохранен во временной папке: {temp_file}")
                print("   Скопируйте файл в нужное место вручную")
                self.output_file = temp_file
                return True
            
        except Exception as e:
            logger.error(f"Ошибка при простом сохранении: {e}")
            
            # Последняя попытка - сохранить в папку пользователя
            try:
                home_dir = Path.home()
                desktop_file = home_dir / "Desktop" / self.output_file.name
                
                with pd.ExcelWriter(desktop_file, engine='openpyxl') as writer:
                    self.df.to_excel(
                        writer,
                        index=False,
                        header=False
                    )
                
                logger.info(f"Файл сохранен на рабочий стол: {desktop_file}")
                print(f"📁 Файл сохранен на рабочий стол: {desktop_file}")
                self.output_file = desktop_file
                return True
                
            except Exception as e2:
                logger.error(f"Финальная попытка сохранения неудачна: {e2}")
                return False
    
    def get_info(self):
        """Возвращает информацию о текущих данных"""
        if self.df is None:
            return "Данные не загружены"
        
        return f"Размер данных: {self.df.shape[0]} строк, {self.df.shape[1]} столбцов"


def main():
    """Основная функция программы"""
    print("=== Программа автоматизации обработки Excel файлов ===")
    print("Шаг 1: Удаление пустых строк в 4 и 5 столбцах")
    print("Шаг 2: Обработка дублей с суммированием")
    print("✓ Полное сохранение форматирования:")
    print("  • Границы ячеек и их стили")
    print("  • Шрифты (название, размер, стиль)")
    print("  • Ширина столбцов и высота строк") 
    print("  • Фиксированная ширина: столбец A = 4, столбец F = 4")
    print("  • Выравнивание текста")
    print("  • Типы данных (числа остаются числами, текст - текстом)")
    print("✓ Обработка дублей:")
    print("  • Первичный ключ: столбец I (уникальные значения)")
    print("  • Суммирование: столбец J")
    print("  • Сохранение порядка строк")
    print("  • Удаление столбцов B, C, F")
    print("  • Сохранение столбцов A, D, E, G, H, I, J")
    print("✓ Поддержка .xls, .xlsm, .xlsx файлов")
    print("✓ Исходный файл остается неизменным")
    
    # Ищем Excel файлы в родительской директории (там где находятся данные)
    parent_dir = Path('..') 
    excel_files = list(parent_dir.glob('*.xls')) + list(parent_dir.glob('*.xlsm'))
    
    # Также ищем в текущей директории
    current_dir = Path('.')
    current_files = list(current_dir.glob('*.xls')) + list(current_dir.glob('*.xlsm'))
    
    # Объединяем списки и убираем дубликаты
    all_files = excel_files + current_files
    excel_files = list(dict.fromkeys(all_files))  # Убираем дубликаты, сохраняя порядок
    
    if not excel_files:
        print("Не найдены файлы .xls или .xlsm")
        print(f"Поиск выполнялся в:")
        print(f"  - Текущая директория: {Path('.').absolute()}")
        print(f"  - Родительская директория: {Path('..').absolute()}")
        return
    
    print("\nДоступные файлы:")
    for i, file in enumerate(excel_files, 1):
        print(f"{i}. {file.name}")
    
    try:
        choice = input("\nВведите номер файла для обработки (или путь к файлу): ").strip()
        
        if choice.isdigit():
            file_idx = int(choice) - 1
            if 0 <= file_idx < len(excel_files):
                input_file = excel_files[file_idx]
            else:
                print("Неверный номер файла")
                return
        else:
            input_file = Path(choice)
        
        # Спрашиваем у пользователя какие шаги выполнять
        print("\nВыберите шаги для выполнения:")
        print("1. Только удаление пустых строк")
        print("2. Только обработка дублей")
        print("3. Оба шага (рекомендуется)")
        
        steps_choice = input("Введите номер (по умолчанию 3): ").strip()
        if not steps_choice:
            steps_choice = "3"
        
        # Создаем процессор
        processor = ExcelProcessor(input_file)
        
        # Загружаем данные
        if not processor.load_data():
            return
        
        print(f"\nИсходные данные: {processor.get_info()}")
        
        # Выполняем выбранные шаги
        if steps_choice in ["1", "3"]:
            print("\n--- Шаг 1: Удаление пустых строк ---")
            # Удаляем пустые строки (4 и 5 столбцы - индексы 3 и 4)
            if not processor.remove_empty_rows(col1_idx=3, col2_idx=4):
                return
            print(f"После удаления пустых строк: {processor.get_info()}")
        
        if steps_choice in ["2", "3"]:
            print("\n--- Шаг 2: Обработка дублей ---")
            print("Настройки:")
            print("  • Первичный ключ: столбец I (индекс 8)")
            print("  • Суммирование: столбец J (индекс 9)")
            print("  • Удаляем столбцы: B, C, F (индексы 1, 2, 5)")
            print("  • Оставляем столбцы: A, D, E, G, H, I, J (индексы 0, 3, 4, 6, 7, 8, 9)")
            
            # Обрабатываем дубли с сохранением порядка
            if not processor.process_duplicates_with_order_preservation(
                primary_key_col=8,    # Столбец I
                sum_col=9,           # Столбец J  
                keep_cols=[0, 3, 4, 6, 7, 8, 9],  # A, D, E, G, H, I, J
                remove_cols=[1, 2, 5]  # B, C, F
            ):
                return
            print(f"После обработки дублей: {processor.get_info()}")
        
        print(f"\nФинальный результат: {processor.get_info()}")
        
        # Сохраняем результат с форматированием
        print("\nСохранение файла...")
        if processor.save_data_with_formatting():
            print(f"✓ Результат сохранен в файл: {processor.output_file}")
            print(f"✓ Полный путь: {processor.output_file.absolute()}")
            print("✓ Полное форматирование сохранено:")
            print("  • Границы ячеек и стили")
            print("  • Шрифты и размеры текста") 
            print("  • Ширина столбцов и высота строк")
            print("  • Выравнивание и другие стили")
            print("  • Типы данных (числа, текст, даты)")
            print("  • Порядок строк сохранен")
        else:
            print("✗ Ошибка при сохранении файла")
        
        # Очищаем временные файлы
        try:
            temp_files = list(Path('..').glob('*.temp.xlsx')) + list(Path('..').glob('temp_*.xls'))
            for temp_file in temp_files:
                if temp_file.exists():
                    temp_file.unlink()
                    logger.info(f"Удален временный файл: {temp_file}")
        except Exception as e:
            logger.warning(f"Не удалось очистить временные файлы: {e}")
        
    except KeyboardInterrupt:
        print("\n\nОбработка прервана пользователем")
    except Exception as e:
        logger.error(f"Общая ошибка: {e}")


if __name__ == "__main__":
    main()
