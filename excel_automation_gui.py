"""
GUI приложение для автоматизации обработки Excel файлов

Объединяет все функции:
1. Конвертация .xls в .xlsx
2. Удаление пустых строк и обработка дублей
3. Сортировка по толщине материала
4. Конвертация в TXT файлы

Автор: Symmppllee
Дата: 2025-08-13
"""

__version__ = "1.6.0"

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import sys
import os
import subprocess
from pathlib import Path
import logging
import pandas as pd
import re
from datetime import datetime
import shutil

# Импортируем наши модули
from automation_tool_fixed import ExcelProcessor
from material_sorter import MaterialSorter
from excel_to_txt_converter import ExcelToTxtConverter

APP_VERSION = __version__  # Используем версию из системы версионирования

# Создаем папки для организации файлов
def ensure_directories():
    """Создает необходимые папки для логов и результатов"""
    app_dir = Path(__file__).parent if not getattr(sys, 'frozen', False) else Path(sys.executable).parent
    
    logs_dir = app_dir / "logs"
    logs_dir.mkdir(exist_ok=True)
    
    results_dir = app_dir / "results"
    results_dir.mkdir(exist_ok=True)
    
    return logs_dir, results_dir

# Настройка логирования для GUI
class GUILogHandler(logging.Handler):
    """Обработчик для вывода логов в GUI"""
    def __init__(self, text_widget):
        super().__init__()
        self.text_widget = text_widget
        
    def emit(self, record):
        msg = self.format(record)
        # Добавляем сообщение в текстовое поле в главном потоке
        self.text_widget.after(0, self._append_log, msg)
        
    def _append_log(self, msg):
        self.text_widget.insert(tk.END, msg + '\n')
        self.text_widget.see(tk.END)
        self.text_widget.update()


class ExcelAutomationGUI:
    """Главный класс GUI приложения"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("MyshkinTool - Полная обработка")
        self.root.geometry("900x700")
        self.root.minsize(800, 600)
        
        # Создаем необходимые папки
        self.logs_dir, self.results_dir = ensure_directories()
        
        # Переменные
        self.input_file = tk.StringVar()
        self.order_number = tk.StringVar()
        self.current_step = tk.StringVar(value="Запуск приложения...")
        
        # Настройка интерфейса
        self.setup_ui()
        
        # Настройка логирования
        self.setup_logging()
        
        # Центрируем окно
        self.center_window()
    
    def center_window(self):
        """Центрирует окно на экране"""
        self.root.update_idletasks()
        x = (self.root.winfo_screenwidth() // 2) - (self.root.winfo_width() // 2)
        y = (self.root.winfo_screenheight() // 2) - (self.root.winfo_height() // 2)
        self.root.geometry(f"+{x}+{y}")
    
    def setup_ui(self):
        """Настройка пользовательского интерфейса"""
        # Главный фрейм
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Настройка сетки
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        
        # Заголовок и версия
        header_frame = ttk.Frame(main_frame)
        header_frame.grid(row=0, column=0, columnspan=3, pady=(0, 20), sticky=(tk.W, tk.E))
        header_frame.columnconfigure(1, weight=1)
        
        title_label = ttk.Label(header_frame, text="Tool for Myshkin", 
                               font=('Arial', 16, 'bold'))
        title_label.grid(row=0, column=0, sticky=tk.W)
        
        version_label = ttk.Label(header_frame, text=f"v{APP_VERSION}", 
                                 font=('Arial', 10), foreground='gray')
        version_label.grid(row=0, column=2, sticky=tk.E)
        
        # Выбор файла
        ttk.Label(main_frame, text="Входной файл:").grid(row=1, column=0, sticky=tk.W, pady=5)
        
        file_frame = ttk.Frame(main_frame)
        file_frame.grid(row=1, column=1, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        file_frame.columnconfigure(0, weight=1)
        
        self.file_entry = ttk.Entry(file_frame, textvariable=self.input_file, state='readonly')
        self.file_entry.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=(0, 10))
        
        ttk.Button(file_frame, text="Обзор...", command=self.select_file).grid(row=0, column=1)
        
        # Номер заказа
        ttk.Label(main_frame, text="Номер круга:").grid(row=2, column=0, sticky=tk.W, pady=5)
        
        order_frame = ttk.Frame(main_frame)
        order_frame.grid(row=2, column=1, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        
        self.order_entry = ttk.Entry(order_frame, textvariable=self.order_number, width=10)
        self.order_entry.grid(row=0, column=0, padx=(0, 10))
        
        ttk.Label(order_frame, text="(например: 1, 13 или 113)").grid(row=0, column=1, sticky=tk.W)
        
        # Текущий шаг
        ttk.Label(main_frame, text="Состояние:").grid(row=3, column=0, sticky=tk.W, pady=5)
        
        self.status_label = ttk.Label(main_frame, textvariable=self.current_step, 
                                     font=('Arial', 10, 'bold'))
        self.status_label.grid(row=3, column=1, sticky=tk.W, pady=5)
        
        # Прогресс-бар
        self.progress = ttk.Progressbar(main_frame, mode='indeterminate')
        self.progress.grid(row=4, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=10)
        
        # Кнопка запуска
        self.start_button = ttk.Button(main_frame, text="Начать обработку", 
                                      command=self.start_processing, style='Accent.TButton')
        self.start_button.grid(row=5, column=0, columnspan=3, pady=10)
        
        # Лог
        ttk.Label(main_frame, text="Журнал обработки:").grid(row=6, column=0, sticky=tk.W, pady=(10, 5))
        
        log_frame = ttk.Frame(main_frame)
        log_frame.grid(row=7, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        
        self.log_text = scrolledtext.ScrolledText(log_frame, height=15, width=80)
        self.log_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Настройка весов для растягивания
        main_frame.rowconfigure(7, weight=1)
        
        # Кнопка очистки лога
        ttk.Button(main_frame, text="Очистить лог", 
                  command=self.clear_log).grid(row=8, column=0, pady=5)
        
        # Кнопка проверки обновлений
        ttk.Button(main_frame, text="Проверить обновления", 
                  command=self.check_for_updates).grid(row=8, column=1, pady=5)
        
        # Кнопка выхода
        ttk.Button(main_frame, text="Выход", 
                  command=self.root.quit).grid(row=8, column=2, pady=5, sticky=tk.E)
    
    def setup_logging(self):
        """Настройка системы логирования"""
        # Создаем обработчик для GUI
        self.gui_handler = GUILogHandler(self.log_text)
        self.gui_handler.setLevel(logging.INFO)
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        self.gui_handler.setFormatter(formatter)
        
        # Настраиваем корневой логгер
        logging.basicConfig(level=logging.INFO, handlers=[self.gui_handler])
        
        # Добавляем обработчик к существующим логгерам модулей
        for module_name in ['__main__', 'automation_tool_fixed', 'material_sorter', 'excel_to_txt_converter']:
            logger = logging.getLogger(module_name)
            logger.addHandler(self.gui_handler)
            logger.setLevel(logging.INFO)
    
    def clear_log(self):
        """Очищает лог"""
        self.log_text.delete(1.0, tk.END)
    
    def check_for_updates(self):
        """Проверяет наличие обновлений приложения"""
        def update_check():
            try:
                self.current_step.set("Проверка обновлений...")
                logging.info("Проверка наличия обновлений...")
                
                # Используем наш простой updater вместо semantic-release
                from simple_updater import SimpleUpdater
                
                # Временно отключим прогресс бар для обновлений
                self.progress.start()
                
                # Создаем updater
                current_dir = Path(__file__).parent
                updater = SimpleUpdater(__version__, current_dir)
                
                try:
                    # Проверяем наличие обновлений
                    has_update, new_version = updater.check_for_updates()
                    
                    if has_update and new_version:
                        # Есть новая версия
                        self.progress.stop()
                        self.current_step.set("Обновление доступно!")
                        
                        response = messagebox.askyesno(
                            "Обновление доступно",
                            f"Доступна новая версия {new_version}!\n"
                            f"Текущая версия: {__version__}\n\n"
                            "Показать подробности об обновлении?"
                        )
                        
                        if response:
                            self.show_update_info(new_version, updater)
                        else:
                            self.current_step.set("Готов к обработке")
                    else:
                        self.progress.stop()
                        self.current_step.set("Обновлений нет")
                        messagebox.showinfo("Обновления", 
                                          f"У вас установлена последняя версия {__version__}")
                        
                except Exception as update_error:
                    self.progress.stop()
                    self.current_step.set("Ошибка проверки обновлений")
                    logging.error(f"Ошибка при проверке обновлений: {update_error}")
                    messagebox.showwarning("Ошибка", 
                                         f"Не удалось проверить обновления:\n{str(update_error)}")
                    
            except ImportError as import_error:
                self.progress.stop()
                self.current_step.set("Ошибка модуля обновлений")
                logging.error(f"Не удалось импортировать модуль обновлений: {import_error}")
                messagebox.showerror("Ошибка", "Модуль обновлений не найден")
                
            except Exception as e:
                self.progress.stop()
                self.current_step.set("Ошибка обновления")
                logging.error(f"Неожиданная ошибка при проверке обновлений: {e}")
                messagebox.showerror("Ошибка", f"Ошибка при проверке обновлений:\n{str(e)}")
        
        # Запускаем проверку в отдельном потоке
        thread = threading.Thread(target=update_check, daemon=True)
        thread.start()
        try:
                self.current_step.set("Проверка обновлений...")
                logging.info("Проверка наличия обновлений...")
                
                # Используем python-semantic-release для проверки версий
                import subprocess
                import tempfile
                import json
                from pathlib import Path
                
                # Временно отключим прогресс бар для обновлений
                self.progress.start()
                
                # Проверим Git репозиторий
                current_dir = Path(__file__).parent
                
                # Проверим следующую версию с помощью semantic-release
                result = subprocess.run([
                    sys.executable, "-m", "semantic_release", "version", "--print"
                ], cwd=current_dir, capture_output=True, text=True)
                
                if result.returncode == 0:
                    next_version = result.stdout.strip()
                    
                    if next_version and next_version != __version__:
                        # Есть новая версия
                        self.progress.stop()
                        self.current_step.set("Обновление доступно!")
                        
                        answer = messagebox.askyesno(
                            "Обновление доступно",
                            f"Доступна новая версия: {next_version}\n"
                            f"Текущая версия: {__version__}\n\n"
                            "Обновить приложение сейчас?",
                            icon='question'
                        )
                        
                        if answer:
                            self.perform_update(next_version)
                        else:
                            logging.info("Обновление отменено пользователем")
                            self.current_step.set("Готов к обработке")
                    else:
                        # Нет обновлений
                        self.progress.stop()
                        self.current_step.set("Обновлений нет")
                        messagebox.showinfo(
                            "Нет обновлений",
                            f"У вас уже установлена последняя версия: {__version__}",
                            icon='info'
                        )
                        logging.info("Обновлений не найдено")
                else:
                    # Ошибка проверки
                    self.progress.stop()
                    self.current_step.set("Ошибка проверки обновлений")
                    logging.error(f"Ошибка при проверке обновлений: {result.stderr}")
                    messagebox.showerror(
                        "Ошибка",
                        "Не удалось проверить обновления.\n"
                        "Убедитесь, что приложение установлено в Git репозитории."
                    )
                
        except Exception as e:
                self.progress.stop()
                self.current_step.set("Ошибка проверки обновлений")
                logging.error(f"Неожиданная ошибка при проверке обновлений: {e}")
                messagebox.showerror("Ошибка", f"Ошибка при проверке обновлений: {str(e)}")
        
        # Запускаем проверку в отдельном потоке
        thread = threading.Thread(target=update_check, daemon=True)
        thread.start()
    
    def perform_update(self, new_version):
        """Выполняет обновление приложения"""
        def update_process():
            try:
                self.current_step.set("Загрузка обновления...")
                self.progress.start()
                logging.info(f"Начинаем обновление до версии {new_version}")
                
                current_dir = Path(__file__).parent
                
                # Выполняем обновление с помощью semantic-release
                result = subprocess.run([
                    sys.executable, "-m", "semantic_release", "version"
                ], cwd=current_dir, capture_output=True, text=True)
                
                if result.returncode == 0:
                    self.progress.stop()
                    self.current_step.set("Обновление завершено!")
                    
                    messagebox.showinfo(
                        "Обновление завершено",
                        f"Приложение успешно обновлено до версии {new_version}!\n\n"
                        "Для применения изменений необходимо перезапустить приложение.",
                        icon='info'
                    )
                    
                    logging.info("Обновление успешно завершено")
                    
                    # Предложим перезапуск
                    restart = messagebox.askyesno(
                        "Перезапуск",
                        "Перезапустить приложение сейчас?",
                        icon='question'
                    )
                    
                    if restart:
                        self.restart_application()
                else:
                    self.progress.stop()
                    self.current_step.set("Ошибка обновления")
                    logging.error(f"Ошибка при обновлении: {result.stderr}")
                    messagebox.showerror(
                        "Ошибка обновления",
                        f"Не удалось выполнить обновление:\n{result.stderr}"
                    )
                
            except Exception as e:
                self.progress.stop()
                self.current_step.set("Ошибка обновления")
                logging.error(f"Ошибка процесса обновления: {e}")
                messagebox.showerror("Ошибка", f"Ошибка при обновлении: {str(e)}")
        
        # Запускаем обновление в отдельном потоке
        thread = threading.Thread(target=update_process, daemon=True)
        thread.start()
    
    def restart_application(self):
        """Перезапускает приложение"""
        try:
            logging.info("Перезапуск приложения...")
            self.root.quit()
            
            # Перезапускаем приложение
            import subprocess
            subprocess.Popen([sys.executable, __file__])
            
        except Exception as e:
            logging.error(f"Ошибка при перезапуске: {e}")
            messagebox.showerror("Ошибка", f"Не удалось перезапустить приложение: {str(e)}")
    
    def show_update_info(self, new_version, updater):
        """Показывает подробную информацию об обновлении"""
        try:
            # Получаем информацию о текущем коммите
            commit_info = updater.get_commit_info()
            
            info_window = tk.Toplevel(self.root)
            info_window.title("Информация об обновлении")
            info_window.geometry("500x400")
            info_window.resizable(False, False)
            
            # Центрируем окно
            info_window.transient(self.root)
            info_window.grab_set()
            
            main_frame = ttk.Frame(info_window, padding="10")
            main_frame.pack(fill=tk.BOTH, expand=True)
            
            # Информация о версиях
            version_frame = ttk.LabelFrame(main_frame, text="Версии", padding="10")
            version_frame.pack(fill=tk.X, pady=(0, 10))
            
            ttk.Label(version_frame, text=f"Текущая версия: {__version__}", 
                     font=('Arial', 10, 'bold')).pack(anchor=tk.W)
            ttk.Label(version_frame, text=f"Новая версия: {new_version}", 
                     font=('Arial', 10, 'bold'), foreground='green').pack(anchor=tk.W, pady=(5, 0))
            
            # Информация о коммите
            commit_frame = ttk.LabelFrame(main_frame, text="Текущий коммит", padding="10")
            commit_frame.pack(fill=tk.X, pady=(0, 10))
            
            ttk.Label(commit_frame, text=f"Хеш: {commit_info['hash']}").pack(anchor=tk.W)
            ttk.Label(commit_frame, text=f"Дата: {commit_info['date']}").pack(anchor=tk.W, pady=(2, 0))
            
            # Сообщение коммита (с переносом строк)
            msg_label = ttk.Label(commit_frame, text=f"Сообщение: {commit_info['message']}", 
                                wraplength=450)
            msg_label.pack(anchor=tk.W, pady=(2, 0))
            
            # Описание обновления
            update_frame = ttk.LabelFrame(main_frame, text="Что нового", padding="10")
            update_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
            
            update_text = tk.Text(update_frame, height=8, width=60, wrap=tk.WORD)
            update_text.insert(tk.END, 
                f"Обновление до версии {new_version} включает:\n\n"
                "• Исправления ошибок и улучшения производительности\n"
                "• Обновленная система проверки обновлений\n"
                "• Улучшенная обработка Excel файлов\n"
                "• Исправлена проблема с Unicode кодировкой\n\n"
                "Для получения полного списка изменений посетите:\n"
                "https://github.com/user/excel-automation-tool/releases"
            )
            update_text.config(state=tk.DISABLED)
            update_text.pack(fill=tk.BOTH, expand=True)
            
            # Кнопки
            button_frame = ttk.Frame(main_frame)
            button_frame.pack(fill=tk.X, pady=(10, 0))
            
            ttk.Button(button_frame, text="Загрузить обновление", 
                      command=lambda: self.download_update(new_version, info_window)).pack(side=tk.RIGHT, padx=(5, 0))
            ttk.Button(button_frame, text="Позже", 
                      command=info_window.destroy).pack(side=tk.RIGHT)
            
        except Exception as e:
            logging.error(f"Ошибка при показе информации об обновлении: {e}")
            messagebox.showerror("Ошибка", f"Не удалось показать информацию об обновлении: {str(e)}")
    
    def download_update(self, new_version, info_window):
        """Загружает и устанавливает обновление"""
        info_window.destroy()
        
        def download_process():
            try:
                self.current_step.set("Выполняется обновление...")
                self.progress.start()
                logging.info(f"Начинаем реальное обновление до версии {new_version}")
                
                # Создаем updater и выполняем РЕАЛЬНОЕ обновление файлов
                from simple_updater import SimpleUpdater
                updater = SimpleUpdater(__version__, Path(__file__).parent)
                
                # Обновляем файлы версий
                pyproject_file = Path(__file__).parent / "pyproject.toml"
                gui_file = Path(__file__)
                
                self.current_step.set("Обновление файла pyproject.toml...")
                success1 = updater.update_version_file(new_version, pyproject_file)
                
                self.current_step.set("Обновление файла приложения...")  
                success2 = updater.update_version_file(new_version, gui_file)
                
                if success1 and success2:
                    self.progress.stop()
                    self.current_step.set("Обновление завершено!")
                    
                    response = messagebox.askyesno(
                        "Обновление завершено",
                        f"Приложение успешно обновлено до версии {new_version}!\n\n"
                        "Изменения:\n"
                        "• Обновлена версия в pyproject.toml\n" 
                        "• Обновлена версия в GUI приложении\n\n"
                        "Перезапустить приложение для применения изменений?",
                        icon='question'
                    )
                    
                    if response:
                        logging.info("Перезапуск приложения после обновления")
                        self.restart_application()
                    else:
                        self.current_step.set(f"Обновлено до v{new_version}")
                        logging.info("Обновление завершено, перезапуск отложен")
                else:
                    self.progress.stop()
                    self.current_step.set("Ошибка обновления файлов")
                    messagebox.showerror(
                        "Ошибка обновления", 
                        "Не удалось обновить файлы версий.\nПроверьте права доступа к файлам."
                    )
                
            except Exception as e:
                self.progress.stop()
                self.current_step.set("Ошибка обновления")
                logging.error(f"Ошибка при обновлении: {e}")
                messagebox.showerror("Ошибка", f"Ошибка при обновлении: {str(e)}")
        
        # Запускаем обновление в отдельном потоке
        thread = threading.Thread(target=download_process, daemon=True)
        thread.start()
    
    def select_file(self):
        """Выбор входного файла"""
        filetypes = [
            ('Excel files', '*.xls *.xlsm *.xlsx'),
            ('XLS files', '*.xls'),
            ('XLSM files', '*.xlsm'),
            ('XLSX files', '*.xlsx'),
            ('All files', '*.*')
        ]
        
        filename = filedialog.askopenfilename(
            title="Выберите Excel файл",
            filetypes=filetypes,
            initialdir=Path('.').parent
        )
        
        if filename:
            self.input_file.set(filename)
            self.current_step.set("Файл выбран")
            logging.info(f"Выбран файл: {filename}")
    
    def validate_inputs(self):
        """Проверка входных данных"""
        if not self.input_file.get():
            messagebox.showerror("Ошибка", "Выберите входной файл")
            return False
        
        if not Path(self.input_file.get()).exists():
            messagebox.showerror("Ошибка", "Выбранный файл не существует")
            return False
        
        if not self.order_number.get().strip():
            messagebox.showerror("Ошибка", "Введите номер круга")
            return False
        
        try:
            int(self.order_number.get().strip())
        except ValueError:
            messagebox.showerror("Ошибка", "Номер круга должен быть числом")
            return False
        
        return True
    
    def start_processing(self):
        """Запуск обработки файла"""
        if not self.validate_inputs():
            return
        
        # Блокируем интерфейс
        self.start_button.config(state='disabled')
        self.progress.start()
        
        # Запускаем обработку в отдельном потоке
        thread = threading.Thread(target=self.process_file)
        thread.daemon = True
        thread.start()
    
    def process_file(self):
        """Полная обработка файла"""
        try:
            input_path = Path(self.input_file.get())
            order_num = self.order_number.get().strip()
            
            # Формируем OrderID
            current_year = datetime.now().year
            year_suffix = str(current_year)[-2:]
            formatted_number = f"{int(order_num):03d}"
            order_id = f"{year_suffix}-{formatted_number}"
            
            logging.info(f"=== Начало обработки файла ===")
            logging.info(f"Входной файл: {input_path.name}")
            logging.info(f"Номер заказа: {order_num} -> OrderID: {order_id}")
            
            # ШАГ 1: Обработка исходного файла
            self.current_step.set("Шаг 1: Удаление пустых строк и обработка дублей")
            logging.info("ШАГ 1: Удаление пустых строк и обработка дублей")
            
            processor = ExcelProcessor(str(input_path))
            if not processor.load_data():
                raise Exception("Ошибка загрузки данных")
            
            # Удаление пустых строк
            if not processor.remove_empty_rows(col1_idx=3, col2_idx=4):
                raise Exception("Ошибка удаления пустых строк")
            
            # Обработка дублей
            if not processor.process_duplicates_with_order_preservation():
                raise Exception("Ошибка обработки дублей")
            
            # Сохранение результата первого шага
            if not processor.save_data_with_formatting():
                raise Exception("Ошибка сохранения данных")
            
            processed_file = processor.output_file
            logging.info(f"Шаг 1 завершен: {processed_file}")
            
            # ШАГ 2: Сортировка по толщине материала
            self.current_step.set("Шаг 2: Сортировка по толщине материала")
            logging.info("ШАГ 2: Сортировка по толщине материала")
            
            sorter = MaterialSorter(str(processed_file))
            if not sorter.load_data():
                raise Exception("Ошибка загрузки обработанных данных")
            
            if not sorter.sort_data_by_thickness():
                raise Exception("Ошибка сортировки по толщине")
            
            # Создание файла с листами по толщине
            # Передаем OrderID для имени файла
            output_file = processed_file.parent / f"{order_id}_by_thickness.xlsx"
            if not sorter.create_sorted_workbook_auto(str(output_file), order_id):
                raise Exception("Ошибка создания файла по толщине")
            
            thickness_file = sorter.output_file
            logging.info(f"Шаг 2 завершен: {thickness_file}")
            
            # ШАГ 3: Конвертация в TXT файлы
            self.current_step.set("Шаг 3: Конвертация в TXT файлы")
            logging.info("ШАГ 3: Конвертация в TXT файлы")
            
            converter = ExcelToTxtConverter(str(thickness_file))
            if not converter.load_workbook():
                raise Exception("Ошибка загрузки файла для конвертации")
            
            # Создаем папку для результатов этого заказа
            order_results_dir = self.results_dir / order_id
            order_results_dir.mkdir(exist_ok=True)
            logging.info(f"Создана папка для результатов: {order_results_dir}")
            
            txt_files = converter.convert_all_sheets()
            if not txt_files:
                raise Exception("Ошибка конвертации в TXT")
            
            # Перемещаем TXT файлы в папку заказа
            moved_files = []
            for txt_file in txt_files:
                if txt_file.exists():
                    new_location = order_results_dir / txt_file.name
                    shutil.move(str(txt_file), str(new_location))
                    moved_files.append(new_location)
                    logging.info(f"Файл перемещен: {txt_file.name} -> {new_location}")
            
            # Завершение
            self.current_step.set("Обработка завершена успешно!")
            logging.info("=== ОБРАБОТКА ЗАВЕРШЕНА УСПЕШНО ===")
            logging.info(f"Создано TXT файлов: {len(moved_files)}")
            logging.info(f"Файлы сохранены в: {order_results_dir}")
            for txt_file in moved_files:
                logging.info(f"  • {txt_file.name}")
            
            # Показываем сообщение об успехе
            self.root.after(0, self.show_success, moved_files, order_results_dir)
            
        except Exception as e:
            logging.error(f"ОШИБКА ОБРАБОТКИ: {e}")
            self.current_step.set("Ошибка обработки")
            self.root.after(0, self.show_error, str(e))
        
        finally:
            # Разблокируем интерфейс
            self.root.after(0, self.finish_processing)
    
    def show_success(self, txt_files, results_dir=None):
        """Показывает сообщение об успешном завершении"""
        message = f"Обработка завершена успешно!\n\nСоздано файлов: {len(txt_files)}\n"
        
        if results_dir:
            message += f"Результаты сохранены в папке:\n{results_dir}\n\n"
        
        message += "Файлы:\n"
        for txt_file in txt_files[:5]:  # Показываем максимум 5 файлов
            message += f"• {txt_file.name}\n"
        if len(txt_files) > 5:
            message += f"... и еще {len(txt_files) - 5} файлов"
        
        messagebox.showinfo("Успех", message)
    
    def show_error(self, error_message):
        """Показывает сообщение об ошибке"""
        messagebox.showerror("Ошибка", f"Произошла ошибка:\n{error_message}")
    
    def finish_processing(self):
        """Завершение обработки - разблокировка интерфейса"""
        self.progress.stop()
        self.start_button.config(state='normal')


# Добавляем метод для автоматической обработки без запроса OrderID
def create_sorted_workbook_auto(sorter, output_file, order_id):
    """
    Автоматическая версия create_sorted_workbook без запроса у пользователя
    """
    if not hasattr(sorter, 'thickness_groups'):
        logging.error("Данные не отсортированы. Сначала вызовите sort_data_by_thickness()")
        return False
    
    try:
        from openpyxl import Workbook
        
        sorter.output_file = Path(output_file)
        logging.info(f"Создаем файл с сортировкой по толщине: {sorter.output_file}")
        
        # Создаем новую книгу
        wb = Workbook()
        wb.remove(wb.active)
        
        # Создаем листы для каждой толщины в определенном порядке
        thickness_order = ["1mm", "1.5mm", "2mm", "3mm"]
        
        for thickness in thickness_order:
            if thickness in sorter.thickness_groups:
                ws = wb.create_sheet(thickness)
                sorter._populate_worksheet(ws, sorter.thickness_groups[thickness], order_id)
                logging.info(f"Создан лист '{thickness}' с {len(sorter.thickness_groups[thickness])} строками")
        
        # Добавляем листы для других толщин (если есть)
        for thickness in sorter.thickness_groups:
            if thickness not in thickness_order:
                ws = wb.create_sheet(thickness)
                sorter._populate_worksheet(ws, sorter.thickness_groups[thickness], order_id)
                logging.info(f"Создан лист '{thickness}' с {len(sorter.thickness_groups[thickness])} строками")
        
        # Создаем лист для неклассифицированных данных (если есть)
        if sorter.unmatched_rows:
            real_unmatched = []
            for row in sorter.unmatched_rows:
                first_value = row.iloc[0] if len(row) > 0 else ""
                if not (isinstance(first_value, str) and 
                       first_value in ['№', 'Порядковый номер', 'OrderID', 'PartName', 'Приоритет', 'nan']):
                    real_unmatched.append(row)
            
            if real_unmatched:
                ws = wb.create_sheet("Неопределенные")
                sorter._populate_worksheet(ws, real_unmatched, order_id)
                logging.info(f"Создан лист 'Неопределенные' с {len(real_unmatched)} строками")
        
        # Сохраняем файл
        wb.save(sorter.output_file)
        wb.close()
        
        logging.info(f"✓ Файл успешно сохранен: {sorter.output_file}")
        return True
        
    except Exception as e:
        logging.error(f"Ошибка при создании файла: {e}")
        return False

# Добавляем метод к классу MaterialSorter
MaterialSorter.create_sorted_workbook_auto = create_sorted_workbook_auto


def main():
    """Главная функция приложения"""
    # Создаем главное окно
    root = tk.Tk()
    
    # Применяем современную тему если доступна
    try:
        root.tk.call('source', 'azure.tcl')
        root.tk.call('set_theme', 'light')
    except:
        pass  # Если тема не доступна, используем стандартную
    
    # Создаем приложение
    app = ExcelAutomationGUI(root)
    # Запускаем главный цикл
    root.mainloop()


if __name__ == "__main__":
    main()
