"""
Модуль для конвертации Excel файлов в TXT формат

Этот модуль берет файл Excel с листами по толщине материала
и создает TXT файлы для каждого листа.

Автор: Автоматизация обработки Excel
Дата: 2025-08-12
"""

import pandas as pd
import re
import sys
from pathlib import Path
from openpyxl import load_workbook
import logging

# Настройка логирования
def setup_logging():
    """Настраивает логирование в папку logs"""
    # Определяем папку приложения
    if getattr(sys, 'frozen', False):
        app_dir = Path(sys.executable).parent
    else:
        app_dir = Path(__file__).parent
    
    # Создаем папку logs
    logs_dir = app_dir / "logs"
    logs_dir.mkdir(exist_ok=True)
    
    # Настраиваем логирование
    log_file = logs_dir / 'excel_to_txt.log'
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler()
        ]
    )
    return logging.getLogger(__name__)

logger = setup_logging()


class ExcelToTxtConverter:
    """Класс для конвертации Excel файлов в TXT формат"""
    
    def __init__(self, input_file: str):
        """
        Инициализация конвертера
        
        Args:
            input_file (str): Путь к входному Excel файлу с листами по толщине
        """
        self.input_file = Path(input_file)
        self.output_dir = None
        self.workbook = None
        
        # Проверяем существование файла
        if not self.input_file.exists():
            raise FileNotFoundError(f"Файл {input_file} не найден")
        
        # Проверяем расширение
        if self.input_file.suffix.lower() not in ['.xlsx', '.xlsm']:
            raise ValueError(f"Неподдерживаемый формат файла: {self.input_file.suffix}")
    
    def load_workbook(self):
        """Загрузка Excel файла"""
        try:
            logger.info(f"Загружаем Excel файл: {self.input_file}")
            self.workbook = load_workbook(self.input_file, data_only=True)
            logger.info(f"Найдены листы: {self.workbook.sheetnames}")
            return True
            
        except Exception as e:
            logger.error(f"Ошибка при загрузке файла: {e}")
            return False
    
    def extract_order_id_from_filename(self, filename: str):
        """
        Извлекает OrderID из имени файла
        
        Args:
            filename (str): Имя файла
            
        Returns:
            str: OrderID или None если не найден
        """
        # Сначала ищем паттерн типа "25-072" (новый формат)
        pattern_new = r'(\d{2}-\d{3})'
        match_new = re.search(pattern_new, filename)
        
        if match_new:
            return match_new.group(1)
        
        # Если не найден, то пытаемся найти старый формат и преобразовать
        # Ищем число в начале имени файла (например "72" из "72.temp_original")
        pattern_old = r'^(\d+)'
        match_old = re.search(pattern_old, filename)
        
        if match_old:
            # Получаем текущий год и берем последние две цифры
            from datetime import datetime
            current_year = datetime.now().year
            year_suffix = str(current_year)[-2:]  # "25" для 2025
            
            # Преобразуем номер в трёхзначный формат
            old_number = int(match_old.group(1))
            formatted_number = f"{old_number:03d}"
            new_order_id = f"{year_suffix}-{formatted_number}"
            
            logger.info(f"Преобразуем старый номер '{match_old.group(1)}' в новый OrderID '{new_order_id}'")
            return new_order_id
        
        return None
    
    def format_sheet_name_for_filename(self, sheet_name: str):
        """
        Форматирует название листа для имени файла
        
        Args:
            sheet_name (str): Название листа (например, "1.5mm", "1mm", "2mm", "3mm")
            
        Returns:
            str: Отформатированное название (например, "15mm", "1mm", "2mm", "3mm")
        """
        # Для листа "1.5mm" убираем точку -> "15mm"
        if sheet_name == "1.5mm":
            return "15mm"
        
        # Для остальных листов оставляем как есть
        return sheet_name
    
    def convert_sheet_to_txt(self, sheet_name: str, output_dir: Path = None):
        """
        Конвертирует один лист Excel в TXT файл
        
        Args:
            sheet_name (str): Название листа
            output_dir (Path): Директория для сохранения (по умолчанию рядом с исходным файлом)
        """
        try:
            if self.workbook is None:
                logger.error("Workbook не загружен. Сначала вызовите load_workbook()")
                return False
            
            if sheet_name not in self.workbook.sheetnames:
                logger.error(f"Лист '{sheet_name}' не найден в файле")
                return False
            
            # Определяем директорию вывода
            if output_dir is None:
                output_dir = self.input_file.parent
            
            # Получаем лист
            worksheet = self.workbook[sheet_name]
            
            # Извлекаем OrderID из имени файла
            order_id = self.extract_order_id_from_filename(self.input_file.name)
            if not order_id:
                logger.warning(f"Не удалось извлечь OrderID из имени файла {self.input_file.name}")
                order_id = "UNKNOWN"
            
            # Форматируем название листа для имени файла
            formatted_sheet_name = self.format_sheet_name_for_filename(sheet_name)
            
            # Создаем имя файла: OrderID_толщина.txt (например, "25-072_15mm.txt")
            txt_filename = f"{order_id}_{formatted_sheet_name}.txt"
            txt_filepath = output_dir / txt_filename
            
            logger.info(f"Конвертируем лист '{sheet_name}' в файл '{txt_filename}'")
            
            # Читаем все данные с листа
            data_rows = []
            
            # Получаем все строки из листа
            for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), 1):
                # Преобразуем None в пустые строки и все значения в строки
                converted_row = []
                for cell_value in row:
                    if cell_value is None:
                        converted_row.append("")
                    else:
                        converted_row.append(str(cell_value))
                
                data_rows.append(converted_row)
                
                if row_idx <= 3:  # Логируем первые несколько строк для отладки
                    logger.debug(f"Строка {row_idx}: {converted_row[:5]}...")  # Показываем первые 5 значений
            
            logger.info(f"Прочитано строк с листа '{sheet_name}': {len(data_rows)}")
            
            # Записываем в TXT файл (разделители - табы)
            with open(txt_filepath, 'w', encoding='utf-8') as f:
                for row in data_rows:
                    # Объединяем ячейки через табуляцию
                    txt_line = '\t'.join(row)
                    f.write(txt_line + '\n')
            
            logger.info(f"✓ Файл сохранен: {txt_filepath}")
            return txt_filepath
            
        except Exception as e:
            logger.error(f"Ошибка при конвертации листа '{sheet_name}': {e}")
            return False
    
    def convert_all_sheets(self, output_dir: Path = None):
        """
        Конвертирует все листы Excel файла в отдельные TXT файлы
        
        Args:
            output_dir (Path): Директория для сохранения файлов
            
        Returns:
            list: Список путей к созданным TXT файлам
        """
        if self.workbook is None:
            logger.error("Workbook не загружен. Сначала вызовите load_workbook()")
            return []
        
        try:
            # Определяем директорию вывода
            if output_dir is None:
                output_dir = self.input_file.parent
            else:
                output_dir = Path(output_dir)
            
            self.output_dir = output_dir
            
            # Создаем директорию если не существует
            output_dir.mkdir(parents=True, exist_ok=True)
            
            logger.info(f"Конвертируем все листы в директорию: {output_dir}")
            
            created_files = []
            
            # Конвертируем каждый лист
            for sheet_name in self.workbook.sheetnames:
                logger.info(f"Обрабатываем лист: {sheet_name}")
                
                result = self.convert_sheet_to_txt(sheet_name, output_dir)
                if result:
                    created_files.append(result)
                    logger.info(f"✓ Лист '{sheet_name}' конвертирован")
                else:
                    logger.error(f"✗ Ошибка при конвертации листа '{sheet_name}'")
            
            logger.info(f"Конвертация завершена. Создано файлов: {len(created_files)}")
            return created_files
            
        except Exception as e:
            logger.error(f"Ошибка при конвертации всех листов: {e}")
            return []
    
    def get_info(self):
        """Возвращает информацию о файле и листах"""
        if self.workbook is None:
            return "Файл не загружен"
        
        info = f"Excel файл: {self.input_file.name}\n"
        info += f"Количество листов: {len(self.workbook.sheetnames)}\n"
        info += f"Листы:\n"
        
        for sheet_name in self.workbook.sheetnames:
            worksheet = self.workbook[sheet_name]
            row_count = worksheet.max_row
            col_count = worksheet.max_column
            info += f"  • {sheet_name}: {row_count} строк, {col_count} столбцов\n"
        
        return info.strip()


def main():
    """Основная функция программы конвертации в TXT"""
    print("=== Конвертер Excel в TXT ===")
    print("Конвертирует каждый лист Excel файла в отдельный TXT файл")
    print("✓ Табулированный формат")
    print("✓ Сохранение всех данных")
    print("✓ Автоматическое именование файлов")
    
    # Ищем файлы с листами по толщине в текущей директории
    current_dir = Path('.')
    thickness_files = list(current_dir.glob('*_by_thickness.xlsx'))
    
    if not thickness_files:
        print("Не найдены файлы с листами по толщине (*_by_thickness.xlsx)")
        print(f"Поиск выполнялся в: {current_dir.absolute()}")
        return
    
    print("\nДоступные файлы:")
    for i, file in enumerate(thickness_files, 1):
        print(f"{i}. {file.name}")
    
    try:
        choice = input("\nВведите номер файла для конвертации (или путь к файлу): ").strip()
        
        if choice.isdigit():
            file_idx = int(choice) - 1
            if 0 <= file_idx < len(thickness_files):
                input_file = thickness_files[file_idx]
            else:
                print("Неверный номер файла")
                return
        else:
            input_file = Path(choice)
        
        # Создаем конвертер
        converter = ExcelToTxtConverter(input_file)
        
        # Загружаем файл
        if not converter.load_workbook():
            return
        
        print(f"\n{converter.get_info()}")
        
        # Конвертируем все листы
        print("\nКонвертация в TXT файлы...")
        created_files = converter.convert_all_sheets()
        
        if created_files:
            print(f"\n✓ Конвертация завершена успешно!")
            print(f"✓ Создано файлов: {len(created_files)}")
            print(f"✓ Файлы сохранены в: {converter.output_dir}")
            print("\nСозданные файлы:")
            for file_path in created_files:
                print(f"  • {file_path.name}")
        else:
            print("✗ Ошибка при конвертации")
        
    except KeyboardInterrupt:
        print("\n\nОбработка прервана пользователем")
    except Exception as e:
        logger.error(f"Общая ошибка: {e}")


if __name__ == "__main__":
    main()
