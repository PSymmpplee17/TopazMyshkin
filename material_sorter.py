"""
Модуль для сортировки данных Excel по толщине материала

Этот модуль берет обработанный файл Excel и создает новый файл
с отдельными листами для каждой толщины материала.

Автор: Автоматизация обработки Excel
Дата: 2025-08-12
"""

import pandas as pd
import re
import sys
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import logging

# Настройка логирования
def setup_logging():
    """Настраивает логирование в папку logs"""
    # Определяем папку приложения
    if getattr(sys, 'frozen', False):
        app_dir = Path(sys.executable).parent
    else:
        app_dir = Path(__file__).parent
    
    # Создаем папку logs
    logs_dir = app_dir / "logs"
    logs_dir.mkdir(exist_ok=True)
    
    # Настраиваем логирование
    log_file = logs_dir / 'material_sorting.log'
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler()
        ]
    )
    return logging.getLogger(__name__)

logger = setup_logging()


class MaterialSorter:
    """Класс для сортировки материалов по толщине"""
    
    # Константы для фиксированной ширины столбцов (такие же как в основной программе)
    FIXED_COLUMN_WIDTHS = {
        'A': 4,   # 33 пикселя
        'B': 64,  # 450 пикселей 
        'C': 22,  # 159 пикселей
        'D': 11,  # 81 пиксель
        'E': 10,   # 75 пикселей
        'F': 26,   # 195 пикселей
        'G': 6     # 45 пикселей
    }
    
    # Константы для числовых столбцов
    NUMERIC_COLUMNS = {'G'}  # Столбец G содержит числовые данные (количество)
    
    def __init__(self, input_file: str):
        """
        Инициализация сортировщика материалов
        
        Args:
            input_file (str): Путь к входному файлу (результат основной обработки)
        """
        self.input_file = Path(input_file)
        self.output_file = None
        self.df = None
        
        # Проверяем существование файла
        if not self.input_file.exists():
            raise FileNotFoundError(f"Файл {input_file} не найден")
        
        # Проверяем расширение
        if self.input_file.suffix.lower() not in ['.xlsx', '.xlsm']:
            raise ValueError(f"Неподдерживаемый формат файла: {self.input_file.suffix}")
    
    def load_data(self):
        """Загрузка данных из обработанного Excel файла"""
        try:
            logger.info(f"Загружаем данные из файла: {self.input_file}")
            
            # Загружаем данные
            self.df = pd.read_excel(
                self.input_file,
                engine='openpyxl',
                header=None  # Не используем первую строку как заголовки
            )
            
            logger.info(f"Данные загружены. Размер: {self.df.shape}")
            return True
            
        except Exception as e:
            logger.error(f"Ошибка при загрузке файла: {e}")
            return False
    
    def extract_thickness_from_material(self, material_description: str):
        """
        Извлекает толщину из описания материала
        
        Args:
            material_description (str): Описание материала
            
        Returns:
            str: Толщина в формате "Xmm" или None если не найдена
        """
        if not material_description or pd.isna(material_description):
            return None
            
        # Ищем паттерн типа "1,0", "1,5", "2,0", "3,0" и т.д. (с десятичной дробью)
        pattern_decimal = r'-([0-9]+,[0-9]+)\s'
        match_decimal = re.search(pattern_decimal, str(material_description))
        
        if match_decimal:
            thickness_str = match_decimal.group(1)
            # Заменяем запятую на точку и конвертируем в float
            thickness_float = float(thickness_str.replace(',', '.'))
            
            # Определяем соответствующий лист
            if thickness_float == 1.0:
                return "1mm"
            elif thickness_float == 1.5:
                return "1.5mm"
            elif thickness_float == 2.0:
                return "2mm"
            elif thickness_float == 3.0:
                return "3mm"
            else:
                # Для других значений создаем название листа
                return f"{thickness_str.replace(',', '.')}mm"
        
        # Ищем паттерн типа "1", "2", "3" и т.д. (целые числа без десятичной части)
        pattern_integer = r'-([0-9]+)\s'
        match_integer = re.search(pattern_integer, str(material_description))
        
        if match_integer:
            thickness_str = match_integer.group(1)
            thickness_int = int(thickness_str)
            
            # Определяем соответствующий лист
            if thickness_int == 1:
                return "1mm"
            elif thickness_int == 2:
                return "2mm"
            elif thickness_int == 3:
                return "3mm"
            else:
                # Для других значений создаем название листа
                return f"{thickness_int}mm"
        
        return None
    
    def sort_data_by_thickness(self):
        """Сортирует данные по толщине материала"""
        if self.df is None:
            logger.error("Данные не загружены. Сначала вызовите load_data()")
            return False
        
        try:
            logger.info("Начинаем сортировку данных по толщине материала...")
            
            # Предполагаем что столбец B (индекс 1) содержит описание материала
            # Столбец G (индекс 6) содержит количество
            material_col_index = 1  # Столбец B
            quantity_col_index = 6  # Столбец G
            
            # Подсчитываем общее количество во входных данных (пропуская заголовки)
            total_input_quantity = 0
            for idx, row in self.df.iterrows():
                if idx == 0:  # Пропускаем первую строку (заголовки)
                    continue
                    
                qty_value = row.iloc[quantity_col_index]
                if pd.notna(qty_value):
                    try:
                        if isinstance(qty_value, str):
                            clean_qty = str(qty_value).strip().replace(',', '.').replace(' ', '')
                            qty_num = int(round(float(clean_qty))) if clean_qty else 0
                        else:
                            qty_num = int(round(float(qty_value)))
                        total_input_quantity += qty_num
                    except (ValueError, TypeError):
                        pass
            
            logger.info(f"Общее количество во входных данных: {total_input_quantity}")
            
            # Создаем словарь для группировки данных по толщине
            thickness_groups = {}
            unmatched_rows = []
            
            # Начинаем цикл со второй строки (индекс 1), пропуская заголовки
            for idx, row in self.df.iterrows():
                if idx == 0:  # Пропускаем первую строку (заголовки)
                    logger.debug(f"Пропускаем строку заголовков: {row.iloc[0]}")
                    continue
                    
                material_desc = row.iloc[material_col_index]
                thickness = self.extract_thickness_from_material(material_desc)
                
                if thickness:
                    if thickness not in thickness_groups:
                        thickness_groups[thickness] = []
                    thickness_groups[thickness].append(row)
                    logger.debug(f"Строка {idx}: '{material_desc}' -> {thickness}")
                else:
                    unmatched_rows.append(row)
                    logger.warning(f"Строка {idx}: не удалось определить толщину для '{material_desc}'")
            
            # Подсчитываем общее количество после группировки
            total_grouped_quantity = 0
            logger.info(f"Группировка завершена:")
            for thickness, rows in thickness_groups.items():
                thickness_quantity = 0
                for row in rows:
                    qty_value = row.iloc[quantity_col_index]
                    if pd.notna(qty_value):
                        try:
                            if isinstance(qty_value, str):
                                clean_qty = str(qty_value).strip().replace(',', '.').replace(' ', '')
                                qty_num = int(round(float(clean_qty))) if clean_qty else 0
                            else:
                                qty_num = int(round(float(qty_value)))
                            thickness_quantity += qty_num
                        except (ValueError, TypeError):
                            pass
                total_grouped_quantity += thickness_quantity
                logger.info(f"  {thickness}: {len(rows)} строк, количество: {thickness_quantity}")
            
            if unmatched_rows:
                unmatched_quantity = 0
                for row in unmatched_rows:
                    qty_value = row.iloc[quantity_col_index]
                    if pd.notna(qty_value):
                        try:
                            if isinstance(qty_value, str):
                                clean_qty = str(qty_value).strip().replace(',', '.').replace(' ', '')
                                qty_num = int(round(float(clean_qty))) if clean_qty else 0
                            else:
                                qty_num = int(round(float(qty_value)))
                            unmatched_quantity += qty_num
                        except (ValueError, TypeError):
                            pass
                total_grouped_quantity += unmatched_quantity
                logger.warning(f"  Не классифицировано: {len(unmatched_rows)} строк, количество: {unmatched_quantity}")
            
            logger.info(f"Общее количество после группировки: {total_grouped_quantity}")
            if total_input_quantity != total_grouped_quantity:
                logger.error(f"ПОТЕРЯ ДАННЫХ! Входное количество: {total_input_quantity}, После группировки: {total_grouped_quantity}")
            else:
                logger.info("✓ Количество данных сохранено корректно")
            
            # Сохраняем результат
            self.thickness_groups = thickness_groups
            self.unmatched_rows = unmatched_rows
            self.total_input_quantity = total_input_quantity
            self.total_grouped_quantity = total_grouped_quantity
            
            return True
            
        except Exception as e:
            logger.error(f"Ошибка при сортировке данных: {e}")
            return False
    
    def create_sorted_workbook(self, output_file: str = None):
        """
        Создает новый Excel файл с листами по толщине
        
        Args:
            output_file (str): Путь к выходному файлу
        """
        if not hasattr(self, 'thickness_groups'):
            logger.error("Данные не отсортированы. Сначала вызовите sort_data_by_thickness()")
            return False
        
        try:
            # Сначала получаем OrderID
            # Получаем текущий год и берем последние две цифры
            from datetime import datetime
            current_year = datetime.now().year
            year_suffix = str(current_year)[-2:]  # Последние 2 цифры года (например, "25" для 2025)
            
            # Запрашиваем номер от пользователя
            raw_order_number = input("Введите номер круга (например, 66 или 1 или 113): ").strip()
            
            # Форматируем номер до трёх цифр с ведущими нулями
            try:
                order_number = int(raw_order_number)
                formatted_number = f"{order_number:03d}"  # Форматируем до 3 цифр с ведущими нулями
                order_id = f"{year_suffix}-{formatted_number}"
                print(f"Сформированный OrderID: {order_id}")
            except ValueError:
                print(f"Ошибка: '{raw_order_number}' не является числом. Используем как есть.")
                order_id = raw_order_number

            # Определяем имя выходного файла с OrderID
            if output_file is None:
                # Создаем имя файла с OrderID: "25-072_by_thickness.xlsx"
                output_file = self.input_file.parent / f"{order_id}_by_thickness.xlsx"
            
            self.output_file = Path(output_file)
            
            logger.info(f"Создаем файл с сортировкой по толщине: {self.output_file}")
            
            # Запрашиваем OrderID один раз для всего файла
            print(f"\n=== Создание файла с листами по толщине ===")
            print(f"Всего листов для создания: {len(self.thickness_groups)}")
            
            # Создаем новую книгу
            wb = Workbook()
            
            # Удаляем стандартный лист
            wb.remove(wb.active)
            
            # Создаем листы для каждой толщины в определенном порядке
            thickness_order = ["1mm", "1.5mm", "2mm", "3mm"]
            
            # Сначала создаем листы в нужном порядке
            for thickness in thickness_order:
                if thickness in self.thickness_groups:
                    ws = wb.create_sheet(thickness)
                    self._populate_worksheet(ws, self.thickness_groups[thickness], order_id)
                    logger.info(f"Создан лист '{thickness}' с {len(self.thickness_groups[thickness])} строками")
            
            # Добавляем листы для других толщин (если есть)
            for thickness in self.thickness_groups:
                if thickness not in thickness_order:
                    ws = wb.create_sheet(thickness)
                    self._populate_worksheet(ws, self.thickness_groups[thickness], order_id)
                    logger.info(f"Создан лист '{thickness}' с {len(self.thickness_groups[thickness])} строками")
            
            # Создаем лист для неклассифицированных данных (только если есть данные, не считая заголовки)
            if self.unmatched_rows:
                # Проверяем, что в неопределенных строках есть реальные данные (не только заголовки)
                real_unmatched = []
                for row in self.unmatched_rows:
                    first_value = row.iloc[0] if len(row) > 0 else ""
                    # Пропускаем заголовки
                    if not (isinstance(first_value, str) and 
                           first_value in ['№', 'Порядковый номер', 'OrderID', 'PartName', 'Приоритет', 'nan']):
                        real_unmatched.append(row)
                
                if real_unmatched:
                    ws = wb.create_sheet("Неопределенные")
                    self._populate_worksheet(ws, real_unmatched, order_id)
                    logger.info(f"Создан лист 'Неопределенные' с {len(real_unmatched)} строками")
                else:
                    logger.info("Неопределенных данных нет (только заголовки)")
            else:
                logger.info("Все данные успешно классифицированы по толщине")
            
            # Сохраняем файл
            wb.save(self.output_file)
            wb.close()
            
            logger.info(f"✓ Файл успешно сохранен: {self.output_file}")
            return True
            
        except Exception as e:
            logger.error(f"Ошибка при создании файла: {e}")
            return False
    
    def _populate_worksheet(self, worksheet, rows_data, order_id):
        """
        Заполняет лист данными в новом формате с 27 столбцами
        
        Args:
            worksheet: Объект листа openpyxl
            rows_data: Список строк данных (pandas Series)
            order_id: OrderID введенный пользователем для всех листов
        """
        try:
            # Получаем сегодняшнюю дату в формате 7/24/2025 (месяц/день/год)
            from datetime import datetime
            today = datetime.now()
            today_date = f"{today.month}/{today.day}/{today.year}"
            
            # Определяем заголовки столбцов
            headers = [
                'OrderID',           # A
                'PartName',          # B
                'QuantityOrdered',   # C
                'QuantityNested',    # D
                'QuantityCompleted', # E
                'ExtraAllowed',      # F
                'Machine',           # G
                'AssemblyID',        # H
                'DueDate',           # I
                'DateWindow',        # J
                'Priority',          # K
                'ForcedPriority',    # L
                'NextPhase',         # M
                'Status',            # N
                'Material',          # O
                'Thickness',         # P
                'AutoTooling',       # Q
                'ScriptTooling',     # R
                'ScriptName',        # S
                'ManualNesting',     # T
                'Drawing',           # U
                'Turret',            # V
                'ProductionLabel',   # W
                'Revision',          # X
                'BendingMode',       # Y
                'BendingParameters', # Z
                'Parameters'         # AA
            ]
            
            # Записываем заголовки в первой строке
            for col_idx, header in enumerate(headers, start=1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = Font(name='Calibri', size=11, bold=True)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # Получаем название листа для столбца Thickness
            sheet_name = worksheet.title
            
            # Определяем машину в зависимости от толщины листа
            machine_name = ""
            if sheet_name in ["1mm", "2mm", "3mm"]:
                machine_name = "A5-25"
            elif sheet_name == "1.5mm":
                machine_name = "E5_TOPAZ"
            
            # Записываем данные построчно начиная со второй строки
            for row_idx, row_series in enumerate(rows_data, start=2):
                # Проверяем, что это не заголовок (первая строка данных может содержать заголовки)
                # Если первый элемент строки похож на заголовок, пропускаем
                first_value = row_series.iloc[0] if len(row_series) > 0 else ""
                if (isinstance(first_value, str) and 
                    first_value in ['nan', 'Порядковый номер', 'OrderID', 'PartName', 'Приоритет']):
                    logger.debug(f"Пропускаем заголовочную строку: {first_value}")
                    continue
                
                # Извлекаем данные из исходных столбцов
                # row_series содержит: A, B, C, D, E, F, G (после обработки automation_tool)
                # Где: A=исх.A, B=исх.D, C=исх.E, D=исх.G, E=исх.H, F=исх.I, G=исх.J
                
                original_a = row_series.iloc[0] if len(row_series) > 0 else ""  # Исх. столбец A
                original_d = row_series.iloc[1] if len(row_series) > 1 else ""  # Исх. столбец D
                original_e = row_series.iloc[2] if len(row_series) > 2 else ""  # Исх. столбец E
                original_g = row_series.iloc[3] if len(row_series) > 3 else ""  # Исх. столбец G
                original_h = row_series.iloc[4] if len(row_series) > 4 else ""  # Исх. столбец H
                original_i = row_series.iloc[5] if len(row_series) > 5 else ""  # Исх. столбец I (обозначение)
                original_j = row_series.iloc[6] if len(row_series) > 6 else 0   # Исх. столбец J (количество)
                
                # Преобразуем обозначение ДСМК в DSMK (БЕЗ удаления суффиксов)
                transformed_designation = ""
                if pd.notna(original_i) and original_i:
                    # Заменяем только ДСМК на DSMK, оставляя все суффиксы
                    transformed_designation = str(original_i).replace('ДСМК.', 'DSMK.')
                    # Убираем только " DXF" в конце, если есть
                    if transformed_designation.endswith(' DXF'):
                        transformed_designation = transformed_designation[:-4]
                
                # Получаем количество как int
                quantity_int = 0
                if pd.notna(original_j):
                    try:
                        if isinstance(original_j, str):
                            clean_qty = str(original_j).strip().replace(',', '.').replace(' ', '')
                            quantity_int = int(round(float(clean_qty))) if clean_qty else 0
                        else:
                            quantity_int = int(round(float(original_j)))
                    except (ValueError, TypeError):
                        quantity_int = 0
                
                # Создаем путь к файлу для столбца U (Drawing)
                drawing_path = ""
                if pd.notna(original_i) and original_i:
                    # Базовый путь
                    base_path = r"\\srvdata\FMS\ncexpress\E5_TOPAZ\PARTDIR"
                    
                    # Берем обозначение из столбца F (transformed_designation)
                    part_name = transformed_designation.strip()
                    
                    # Извлекаем версию DXF из исходного столбца H (в обработанном файле это original_h)
                    version = ""
                    if pd.notna(original_h) and original_h:
                        import re
                        # Ищем цифру в строке типа "3" или "3.0" -> берем первую цифру
                        version_match = re.search(r'(\d+)', str(original_h).strip())
                        if version_match:
                            version_digit = version_match.group(1)
                            version = f"_V{version_digit}"
                        else:
                            version = "_V0"  # По умолчанию, если цифра не найдена
                    else:
                        version = "_V0"  # По умолчанию для пустых значений
                    
                    # Добавляем толщину в зависимости от листа
                    thickness_suffix = ""
                    if sheet_name == "1mm":
                        thickness_suffix = "_1mmZn"
                    elif sheet_name == "1.5mm":
                        thickness_suffix = "_1.5mmZn"
                    elif sheet_name == "2mm":
                        thickness_suffix = "_2mmZn"
                    elif sheet_name == "3mm":
                        thickness_suffix = "_3mmZn"
                    
                    # Собираем полный путь
                    drawing_path = f"{base_path}\\{part_name}{version}{thickness_suffix}"
                    
                    # Создаем полное имя детали для столбца B
                    full_part_name = f"{part_name}{version}{thickness_suffix}"
                
                # Заполняем столбцы согласно новой структуре (используем общий OrderID для всего листа)
                new_row_data = [
                    order_id,                  # A - OrderID (введенный пользователем для всего листа)
                    full_part_name,            # B - PartName (полное имя с версией и толщиной)
                    quantity_int,              # C - QuantityOrdered (количество)
                    0,                         # D - QuantityNested
                    0,                         # E - QuantityCompleted
                    0,                         # F - ExtraAllowed
                    machine_name,              # G - Machine (A5-25 или E5_TOPAZ)
                    "",                        # H - AssemblyID (пустой)
                    today_date,                # I - DueDate (сегодняшняя дата)
                    0,                         # J - DateWindow (в пределах данных заполнен 0)
                    original_g if pd.notna(original_g) else "",  # K - Priority (исх. столбец G)
                    0,                         # L - ForcedPriority
                    0,                         # M - NextPhase
                    0,                         # N - Status
                    "DC01",                    # O - Material
                    f"{float(sheet_name.replace('mm', '')) if sheet_name.replace('mm', '').replace('.', '').isdigit() else 0:.6f}",  # P - Thickness (толщина с 6 знаками после запятой)
                    0,                         # Q - AutoTooling
                    0,                         # R - ScriptTooling
                    "",                        # S - ScriptName (пустой)
                    0,                         # T - ManualNesting (в пределах данных заполнен 0)
                    drawing_path,              # U - Drawing (путь к файлу)
                    "",                        # V - Turret (пустой)
                    "",                        # W - ProductionLabel (пустой)
                    "",                        # X - Revision (пустой)
                    -1,                        # Y - BendingMode
                    "",                        # Z - BendingParameters (пустой)
                    ""                         # AA - Parameters (пустой)
                ]
                
                # Записываем данные в ячейки
                for col_idx, value in enumerate(new_row_data, start=1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    
                    # Применяем форматирование
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    cell.font = Font(name='Calibri', size=11)
                    
                    # Для числовых столбцов устанавливаем правильный формат
                    if col_idx in [3, 4, 5, 6, 12, 13, 14, 17, 18, 25]:  # C, D, E, F, L, M, N, Q, R, Y
                        if isinstance(value, (int, float)):
                            cell.number_format = '0'
            
            # Устанавливаем ширину столбцов
            column_widths = {
                'A': 25, 'B': 25, 'C': 12, 'D': 12, 'E': 12, 'F': 10, 'G': 15,
                'H': 12, 'I': 12, 'J': 12, 'K': 10, 'L': 10, 'M': 10, 'N': 8,
                'O': 10, 'P': 10, 'Q': 10, 'R': 10, 'S': 15, 'T': 20, 'U': 15,
                'V': 10, 'W': 15, 'X': 10, 'Y': 10, 'Z': 15, 'AA': 15
            }
            
            for col_letter, width in column_widths.items():
                worksheet.column_dimensions[col_letter].width = width
            
            logger.info(f"Лист заполнен: {len(rows_data)} строк данных + заголовки")
            
        except Exception as e:
            logger.error(f"Ошибка при заполнении листа: {e}")
            raise
    
    def get_info(self):
        """Возвращает информацию о текущих данных"""
        if not hasattr(self, 'thickness_groups'):
            return "Данные не отсортированы"
        
        info = f"Сортировка по толщине:\n"
        
        # Подсчитываем количество для каждой группы
        for thickness, rows in self.thickness_groups.items():
            thickness_quantity = 0
            for row in rows:
                qty_value = row.iloc[6]  # Столбец G (индекс 6)
                if pd.notna(qty_value):
                    try:
                        if isinstance(qty_value, str):
                            clean_qty = str(qty_value).strip().replace(',', '.').replace(' ', '')
                            qty_num = int(round(float(clean_qty))) if clean_qty else 0
                        else:
                            qty_num = int(round(float(qty_value)))
                        thickness_quantity += qty_num
                    except (ValueError, TypeError):
                        pass
            info += f"  {thickness}: {len(rows)} строк, количество: {thickness_quantity}\n"
        
        if hasattr(self, 'unmatched_rows') and self.unmatched_rows:
            unmatched_quantity = 0
            for row in self.unmatched_rows:
                qty_value = row.iloc[6]  # Столбец G
                if pd.notna(qty_value):
                    try:
                        if isinstance(qty_value, str):
                            clean_qty = str(qty_value).strip().replace(',', '.').replace(' ', '')
                            qty_num = int(round(float(clean_qty))) if clean_qty else 0
                        else:
                            qty_num = int(round(float(qty_value)))
                        unmatched_quantity += qty_num
                    except (ValueError, TypeError):
                        pass
            info += f"  Неопределенные: {len(self.unmatched_rows)} строк, количество: {unmatched_quantity}\n"
        
        # Добавляем общую информацию
        if hasattr(self, 'total_input_quantity') and hasattr(self, 'total_grouped_quantity'):
            info += f"\nОбщее количество:\n"
            info += f"  Входное: {self.total_input_quantity}\n"
            info += f"  После группировки: {self.total_grouped_quantity}\n"
            if self.total_input_quantity != self.total_grouped_quantity:
                info += f"  ⚠️  ПОТЕРЯ ДАННЫХ: {self.total_input_quantity - self.total_grouped_quantity}\n"
            else:
                info += f"  ✓ Данные сохранены корректно\n"
        
        return info.strip()


def main():
    """Основная функция программы сортировки по материалам"""
    print("=== Программа сортировки по толщине материала ===")
    print("Создает листы: 1mm, 1.5mm, 2mm, 3mm")
    print("Анализирует столбец B для определения толщины")
    print("✓ Полное сохранение форматирования")
    print("✓ Числовые форматы сохраняются")
    print("✓ Ширина столбцов как в исходном файле")
    
    # Ищем обработанные файлы в текущей директории
    current_dir = Path('.')
    processed_files = list(current_dir.glob('*_original*.xlsx'))
    
    if not processed_files:
        print("Не найдены обработанные файлы (*_processed*.xlsx)")
        print(f"Поиск выполнялся в: {current_dir.absolute()}")
        return
    
    print("\nДоступные файлы:")
    for i, file in enumerate(processed_files, 1):
        print(f"{i}. {file.name}")
    
    try:
        choice = input("\nВведите номер файла для сортировки (или путь к файлу): ").strip()
        
        if choice.isdigit():
            file_idx = int(choice) - 1
            if 0 <= file_idx < len(processed_files):
                input_file = processed_files[file_idx]
            else:
                print("Неверный номер файла")
                return
        else:
            input_file = Path(choice)
        
        # Создаем сортировщик
        sorter = MaterialSorter(input_file)
        
        # Загружаем данные
        if not sorter.load_data():
            return
        
        print(f"\nИсходные данные: {sorter.df.shape[0]} строк, {sorter.df.shape[1]} столбцов")
        
        # Сортируем по толщине
        print("\n--- Сортировка по толщине материала ---")
        if not sorter.sort_data_by_thickness():
            return
        
        print(f"\n{sorter.get_info()}")
        
        # Создаем новый файл с сортировкой
        print("\nСоздание файла с листами по толщине...")
        if sorter.create_sorted_workbook():
            print(f"✓ Результат сохранен в файл: {sorter.output_file}")
            print(f"✓ Полный путь: {sorter.output_file.absolute()}")
            print("✓ Созданы листы:")
            if hasattr(sorter, 'thickness_groups'):
                for thickness in sorter.thickness_groups.keys():
                    print(f"  • {thickness}")
            if hasattr(sorter, 'unmatched_rows') and sorter.unmatched_rows:
                print(f"  • Неопределенные")
        else:
            print("✗ Ошибка при создании файла")
        
    except KeyboardInterrupt:
        print("\n\nОбработка прервана пользователем")
    except Exception as e:
        logger.error(f"Общая ошибка: {e}")


if __name__ == "__main__":
    main()